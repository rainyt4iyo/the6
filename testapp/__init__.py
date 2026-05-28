from flask import Flask
import pymysql
import openpyxl

CLASSES = ['F1', 'F2', 'F3', 'F4', 'F5M', 'F5W', 'F6M', 'F6W', 'O1', 'O2', 'O3', 'O4M', 'O4W', 'O5M', 'O5W', 'O6M', 'O6W']

app = Flask(__name__)
app.config.from_object('testapp.config')

conn = pymysql.connect(host='localhost',
                       user='t4',
                       password='t4_password',
                       database='the6',
                       cursorclass=pymysql.cursors.DictCursor)
cursor = conn.cursor()


sql = """
CREATE TABLE IF NOT EXISTS player (
    id INT AUTO_INCREMENT PRIMARY KEY, 
    pid INT, 
    class VARCHAR(255), 
    name VARCHAR(255),
    UNIQUE(pid, class)
)
"""
cursor.execute(sql)
conn.commit()

sql = """
CREATE TABLE IF NOT EXISTS result (
    key_id INT AUTO_INCREMENT PRIMARY KEY, 
    pid INT, 
    class VARCHAR(255), 
    kid VARCHAR(255), 
    zt INT,
    UNIQUE(pid, class, kid)
)
"""
cursor.execute(sql)
conn.commit()


print("--------------------------------------")
print("[DB] > 初期化完了")
print("--------------------------------------")

import openpyxl
wb = openpyxl.load_workbook("static.xlsx")

# 2. データの挿入・更新
for class_name in CLASSES:
    if class_name not in wb.sheetnames:
        print(f"シート {class_name} が見つかりません。スキップします。")
        continue
    ws = wb[class_name]
    for i in range(1, 100):
        pid_val = ws.cell(i, 1).value
        name_val = ws.cell(i, 2).value
        
        # 空行を読み込んだ場合はスキップする処理を入れると安全です
        if pid_val is None:
            continue

        # SQLの %s に対応する3つの値を用意する (pid, class, name)
        temp = (pid_val, class_name, name_val)
        
        # 3. pidとclassが重複している場合は name を更新する
        sql = """
        INSERT INTO player (pid, class, name) 
        VALUES (%s, %s, %s) 
        ON DUPLICATE KEY UPDATE name = VALUES(name)
        """
        
        cursor.execute(sql, temp)
        conn.commit()

print("--------------------------------------")



import testapp.views